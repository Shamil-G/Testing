from db_oracle.connect import get_connection
import config as cfg
from openpyxl import load_workbook
import datetime
import os.path

file_name = 'questions.xlsx'
file_path = cfg.REPORTS_PATH + file_name
id_theme = 1


def load_questions():
    now = datetime.datetime.now()

    print("Exist excel file: " + str(os.path.isfile(file_path)) + ' : ' + file_path)
    if not os.path.isfile(file_path):
        return file_name

    wb = load_workbook(file_path)
    print("Книга загружена: " + file_path)
    sheet = wb.active

    print("Подключаем БД")

    con = get_connection()
    cursor = con.cursor()
    cursor.execute("delete from answers a where a.id_question in "
                   "(select id_question from questions q where q.id_theme = " + str(id_theme) + ")")
    cursor.execute("delete from questions where id_theme = " + str(id_theme))
    cursor.execute("commit")
    id_quest = 0
    id_prev_quest = -1
    order_num = 0
    for i in range(2, sheet.max_row):
        id_curr_quest = sheet.cell(row=i, column=2).value
        quest = sheet.cell(row=i, column=3).value
        correctly = sheet.cell(row=i, column=4).value
        answer = sheet.cell(row=i, column=5).value
        order_num = order_num + 1
        if id_curr_quest != id_prev_quest:
            id_quest = id_quest + 1
            order_num = 1
            id_question = cursor.callfunc("admin.add_question", str, [id_theme, id_quest, quest])

        cmd = "insert into answers q (id_answer, id_question, order_num_answer, correctly, active, answer) " \
              "values ( seq_answer.nextval, " + str(id_question) + ", " + str(order_num) + ", '" + correctly + "', 'Y', '" + str(answer) + "')"
        cursor.execute(cmd)
        id_prev_quest = id_curr_quest

    con.commit()
    con.close()
    now = datetime.datetime.now()
    print("Загрузка завершена: " + now.strftime("%d-%m-%Y %H:%M:%S"))
    return file_name


if __name__ == "__main__":
    print("Тестируем загрузку Excel!")
    file_path = 'spool/' + file_name

    load_questions()
